import time
import re , os
import pyautogui
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook

class ChromeDriver:
    # __init__(생성자함수) : ChromeDriver class의 객체(인스턴스) 생성 시, __init__(생성자함수)에 정의 되어있던 코드들이 바로 실행됨
    def set_driver(self) -> any:
        # *-- 인스턴스 변수 정의 *--

        # options 객체 생성
        chrome_options = Options()

        # headless chrome
        chrome_options.add_argument('--headless')

        # 브라우저 꺼짐 방지
        chrome_options.add_experimental_option('detach', True)

        # 불필요한 에러메시지 제거
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        # Service 객체
        service = Service(executable_path=ChromeDriverManager().install())

        # driver 객체
        browser = webdriver.Chrome(service=service, options=chrome_options)
        browser.maximize_window()

        return browser



class Application:
    # 생성자 함수 정의
    def __init__(self):
        # ChromeDriver 객체 생성
        self.chromedriver = ChromeDriver()

        # set_driver 메서드 실행 후 , 리턴 값 멤버변수로 정의
        self.browser = self.chromedriver.set_driver()

        # input_url 메서드의 리턴 값 인스턴스 변수로 정의
        self.url : str = self.input_url()

        # URL 주소의 키워드 변환 메소드의 리턴 값 인스턴스 변수로 정의
        self.keword : str = self.translate_keword()

    # 실행 메서드 정의
    def run(self) -> list :
        # 브라우저 이동
        self.browser.get(url=self.url)
        self.browser.implicitly_wait(15)

        # 리뷰 더보기 클릭
        self.button_click()

        # 리뷰 스크롤 내리기
        self.review_scroll()

        # 데이터 추출하기
        results = self.get_content()

        # 리턴 값 == list 형태
        return results

    def get_content(self) -> list:
        soup = bs(self.browser.page_source, 'html.parser')
        review_list_len = len(soup.select("div.jgIq1 div.RHo1pe"))

        ws_data = []

        for idx in range(review_list_len):
            review_list = soup.select("div.jgIq1 div.RHo1pe")

            # 작성자 이름 가져오기
            name = review_list[idx].select_one("div.X5PpBb")

            # 평점 가져오기
            rating = review_list[idx].select_one("div.iXRFPc")

            # 리뷰 작성일
            review_date = review_list[idx].select_one("span.bp9Aid")

            # 작성내용
            content = review_list[idx].select_one('div.h3YV2d')

            if name == None or name.text == '':
                name = '이름을 알 수 없음'

            else:
                name = name.text.strip()

            if rating == None or re.sub('[가-힣ㄱ-ㅎ.\t5 \n]', '', rating.attrs['aria-label']) == '':
                rating = 0

            else:
                rating = rating.attrs['aria-label']
                rating = int(re.sub('[^0-9]', '', rating))

            if review_date == None == review_date.text:
                review_date = '작성일 알 수 없음'

            else:
                review_date = review_date.text.strip()
                review_date = int(re.sub('[^0-9]', '', review_date).strip())

            if content == None or content.text == '':
                content = '작성된 리뷰내용 없음'

            else:
                content = content.text.strip()

                # 특수문자 때문에 DB에 저장 시 에러가 났었음. 정규표현식으로 특수기호 없애니 정상적으로 들어감
                content = re.sub('[^가-힣ㄱ-ㅎA-Za-z\n \t]', '', content)

            # 엑셀 저장용 데이터 리스트 변수에 저장
            ws_data.append([name, rating, review_date, content])

            print(f"{name}\n{rating}\n{review_date}\n{content}")
            print()

        return ws_data

    # 리뷰 스크롤 내리기 메서드 정의
    def review_scroll(self) -> None:
        count = 0
        while True:
            if count >= 1:
                break

            last_item = self.browser.find_elements(By.CSS_SELECTOR, 'div.RHo1pe')[
                -1]  # 가져온 item 태그(동적)의 제일 마지막 태그(아이템)값 가져오기

            time.sleep(0.5)

            self.browser.execute_script("arguments[0].scrollIntoView(true);", last_item)

            # count 증감
            count += 1

    # 리뷰 더보기 클릭 메서드 정의
    def button_click(self) -> None:
        clk_tag = self.browser.find_elements(By.CSS_SELECTOR, 'div.u4ICaf > div > button')[-1]
        self.browser.execute_script('arguments[0].click()', clk_tag)
        time.sleep(2)


    # 키워드 변환 메서드 정의
    def translate_keword(self) -> str:
        keword = self.url.split('.')
        keword = keword[-2] + '_' + keword[-1]

        return keword

    # return Value 타입 : string
    def input_url(self) -> str:
        os.system('cls')

        url = input('URL 주소를 입력하세요\nEx) https://play.google.com/store/apps/details?id=com.cyworld.minihompy\n:')

        return url


class OpenPyXL:
    def __init__(self):
        # Application 클래스의 인스턴스(객체) 생성 -> OpenPyXL 클래스의 인스턴스 변수로써
        self.app = Application()

        # Application class의 인스턴스 메서드(run)을 실행 후 리턴 되는 값을 인스턴스 변수로 정의 (리턴 된 값은 list 형태로 리턴이 되므로 , 타입형태도 list로 정의)
        self.results : list = self.app.run()

        # Workbook 인스턴스 생성 # 파이썬에서 모든 것은 객체이다
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['이름','평점','작성 날짜','리뷰 내용'])

    def save_file(self) -> None:
        row : int = 2
        for x in self.results:
            self.ws[f'A{row}'] = x[0]
            self.ws[f'B{row}'] = x[1]
            self.ws[f'C{row}'] = x[2]
            self.ws[f'D{row}'] = x[-1]

            row += 1

        # 저장경로 지정
        savePath : str = os.path.abspath('구글스토어_리뷰모음')
        fileName : str = self.app.keword

        if not os.path.exists(savePath):
            os.mkdir(savePath)

        # 파일 저장
        self.wb.save(os.path.join(savePath,fileName))
        self.wb.close()

        # 파일저장 완료 알림
        pyautogui.alert(f'파일 저장완료!\n\n{savePath}')










if __name__ == '__main__' :
    app = OpenPyXL()

    app.save_file()
